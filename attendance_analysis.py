import random
import numpy as np
import pandas as pd
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart.label import DataLabelList


# ── Palette ────────────────────────────────────────────────────────────────────
C_NAVY   = "1B3A6B"
C_TEAL   = "0D7377"
C_GOLD   = "F4A261"
C_CORAL  = "E76F51"
C_LIGHT  = "F0F4F8"
C_WHITE  = "FFFFFF"
C_DARK   = "1A1A2E"
C_GREEN  = "2D6A4F"
C_RED    = "C1121F"
C_GRAY   = "6C757D"
C_LBLUE  = "AED9E0"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_DARK, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color=C_NAVY)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Data generation ────────────────────────────────────────────────────────────
STUDENTS = [
    ("Emma Johnson",   0.92, "excellent"),
    ("Liam Williams",  0.78, "average"),
    ("Sophia Brown",   0.85, "good"),
    ("Noah Jones",     0.61, "poor"),
    ("Olivia Garcia",  0.95, "excellent"),
    ("Mason Davis",    0.70, "average"),
    ("Ava Martinez",   0.88, "good"),
    ("Ethan Wilson",   0.55, "poor"),
    ("Isabella Moore", 0.82, "good"),
    ("Lucas Taylor",   0.91, "excellent"),
    ("Mia Anderson",   0.67, "average"),
    ("Aiden Thomas",   0.74, "average"),
]

SCHOOL_START = date(2024, 9, 2)
SCHOOL_END   = date(2025, 5, 30)

def school_days(start, end):
    days = []
    d = start
    while d <= end:
        if d.weekday() < 5:  # Mon–Fri
            days.append(d)
        d += timedelta(days=1)
    return days

def generate_attendance(students, days):
    random.seed(42)
    records = []
    for name, base_prob, tier in students:
        for d in days:
            # Seasonal / monthly adjustments
            month = d.month
            if month in (12, 1):   adj = -0.08   # Winter illness
            elif month in (3, 4):  adj = -0.04   # Spring absence
            elif month == 5:       adj = +0.03   # End-of-year push
            else:                  adj = 0.0

            # Day-of-week effects
            dow = d.weekday()
            dow_adj = {0: +0.02, 1: +0.01, 2: 0, 3: -0.01, 4: -0.04}[dow]

            p = min(1.0, max(0.0, base_prob + adj + dow_adj))
            present = 1 if random.random() < p else 0
            records.append({
                "student":    name,
                "tier":       tier,
                "date":       d,
                "weekday":    d.strftime("%A"),
                "week":       d.isocalendar()[1],
                "month":      d.month,
                "month_name": d.strftime("%B"),
                "present":    present,
            })
    return pd.DataFrame(records)


# ── Sheet helpers ──────────────────────────────────────────────────────────────
def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def header_row(ws, row, cols, texts, bg=C_NAVY, fg=C_WHITE, size=11, bold=True):
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = font(bold=bold, color=fg, size=size)
        c.fill      = fill(bg)
        c.alignment = align()
        c.border    = border_thin()

def data_cell(ws, row, col, value, bg=C_WHITE, bold=False,
              h="center", number_format=None, color=C_DARK):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font(bold=bold, color=color)
    c.fill      = fill(bg)
    c.alignment = align(h=h)
    c.border    = border_thin()
    if number_format:
        c.number_format = number_format
    return c

def section_title(ws, row, col, text, merge_to_col=None,
                  bg=C_TEAL, size=13):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = font(bold=True, color=C_WHITE, size=size)
    c.fill      = fill(bg)
    c.alignment = align(h="left")
    c.border    = border_medium()
    if merge_to_col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=merge_to_col
        )


# ── Sheet 1: Raw Attendance Log ────────────────────────────────────────────────
def build_raw_log(wb, df):
    ws = wb.create_sheet("📋 Attendance Log")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Banner
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "🏫  Student Attendance Log  —  Academic Year 2024–25"
    c.font      = font(bold=True, color=C_WHITE, size=16)
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="left")

    headers = ["Student", "Attendance Tier", "Date", "Day of Week",
               "Week #", "Month", "Status"]
    header_row(ws, 2, range(1, 8), headers)

    tier_colors = {
        "excellent": ("D1FAE5", "065F46"),
        "good":      ("DBEAFE", "1E40AF"),
        "average":   ("FEF3C7", "92400E"),
        "poor":      ("FEE2E2", "991B1B"),
    }

    for i, row in df.iterrows():
        r = i + 3
        bg  = C_LIGHT if r % 2 == 0 else C_WHITE
        tbg, tfg = tier_colors.get(row.tier, (C_WHITE, C_DARK))

        data_cell(ws, r, 1, row.student,    bg=bg, h="left")
        # Tier badge
        c = ws.cell(row=r, column=2, value=row.tier.capitalize())
        c.font      = font(bold=True, color=tfg)
        c.fill      = fill(tbg)
        c.alignment = align()
        c.border    = border_thin()
        data_cell(ws, r, 3, row.date,        bg=bg, number_format="YYYY-MM-DD")
        data_cell(ws, r, 4, row.weekday,     bg=bg)
        data_cell(ws, r, 5, row.week,        bg=bg)
        data_cell(ws, r, 6, row.month_name,  bg=bg)
        # Status
        label = "✔ Present" if row.present else "✘ Absent"
        sbg   = "D1FAE5" if row.present else "FEE2E2"
        sfg   = C_GREEN  if row.present else C_RED
        c = ws.cell(row=r, column=7, value=label)
        c.font      = font(bold=True, color=sfg)
        c.fill      = fill(sbg)
        c.alignment = align()
        c.border    = border_thin()

    set_col_widths(ws, {"A":22,"B":16,"C":14,"D":14,"E":10,"F":12,"G":14})
    ws.auto_filter.ref = "A2:G2"


# ── Sheet 2: Student Probability Summary ──────────────────────────────────────
def build_summary(wb, df):
    ws = wb.create_sheet("📊 Student Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "🎯  Attendance Probability Summary by Student"
    c.font      = font(bold=True, color=C_WHITE, size=16)
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="left")
    ws.row_dimensions[1].height = 30

    cols = ["Student","Tier","Total Days","Present","Absent",
            "Attend. Prob.","Mon","Tue","Wed","Thu","Fri"]
    header_row(ws, 2, range(1, 12), cols)

    summary = (df.groupby(["student","tier"])
                 .agg(total=("present","count"),
                      present=("present","sum"))
                 .reset_index())
    summary["absent"] = summary["total"] - summary["present"]
    summary["prob"]   = summary["present"] / summary["total"]

    dow_pivot = (df.groupby(["student","weekday"])["present"]
                   .mean().unstack(fill_value=0))
    for day in ["Monday","Tuesday","Wednesday","Thursday","Friday"]:
        if day not in dow_pivot.columns:
            dow_pivot[day] = 0

    tier_order = {"excellent":0,"good":1,"average":2,"poor":3}
    summary["tier_order"] = summary["tier"].map(tier_order)
    summary = summary.sort_values(["tier_order","prob"], ascending=[True,False])

    tier_colors = {
        "excellent": ("D1FAE5","065F46"),
        "good":      ("DBEAFE","1E40AF"),
        "average":   ("FEF3C7","92400E"),
        "poor":      ("FEE2E2","991B1B"),
    }

    for i, (_, row) in enumerate(summary.iterrows()):
        r   = i + 3
        bg  = C_LIGHT if i % 2 == 0 else C_WHITE
        tbg, tfg = tier_colors.get(row.tier, (C_WHITE, C_DARK))

        data_cell(ws, r,  1, row.student,  bg=bg, h="left", bold=True)
        c = ws.cell(row=r, column=2, value=row.tier.capitalize())
        c.font = font(bold=True, color=tfg); c.fill = fill(tbg)
        c.alignment = align(); c.border = border_thin()
        data_cell(ws, r,  3, row.total,   bg=bg)
        data_cell(ws, r,  4, row.present, bg=bg, color=C_GREEN)
        data_cell(ws, r,  5, row.absent,  bg=bg, color=C_RED)

        prob = row.prob
        pbg  = ("D1FAE5" if prob >= .90 else
                "DBEAFE" if prob >= .80 else
                "FEF3C7" if prob >= .70 else "FEE2E2")
        data_cell(ws, r, 6, prob, bg=pbg, bold=True, number_format="0.0%")

        dow_row = dow_pivot.loc[row.student] if row.student in dow_pivot.index else {}
        for j, day in enumerate(["Monday","Tuesday","Wednesday","Thursday","Friday"]):
            v   = dow_row.get(day, 0)
            dbg = ("D1FAE5" if v >= .90 else
                   "DBEAFE" if v >= .80 else
                   "FEF3C7" if v >= .70 else "FEE2E2")
            data_cell(ws, r, 7+j, v, bg=dbg, number_format="0%")

    # Colour-scale on Attend. Prob. column (F3:F14)
    last = 2 + len(summary)
    ws.conditional_formatting.add(
        f"F3:F{last}",
        ColorScaleRule(start_type="num", start_value=0.5, start_color="FEE2E2",
                       mid_type="num",   mid_value=0.75,  mid_color="FEF3C7",
                       end_type="num",   end_value=1.0,   end_color="D1FAE5")
    )

    set_col_widths(ws, {"A":22,"B":13,"C":11,"D":10,"E":10,
                        "F":13,"G":9,"H":9,"I":9,"J":9,"K":9})

    # ── Bar chart: overall attendance probability ──────────────────────────────
    chart = BarChart()
    chart.type        = "bar"
    chart.title       = "Attendance Probability by Student"
    chart.y_axis.title = "Probability"
    chart.x_axis.title = "Student"
    chart.style       = 10
    chart.height      = 14
    chart.width       = 24
    chart.grouping    = "clustered"

    data_ref  = Reference(ws, min_col=6, max_col=6,
                          min_row=2, max_row=2+len(summary))
    cats_ref  = Reference(ws, min_col=1, max_col=1,
                          min_row=3, max_row=2+len(summary))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = C_TEAL
    ws.add_chart(chart, "A18")


# ── Sheet 3: Monthly Trends ────────────────────────────────────────────────────
def build_monthly(wb, df):
    ws = wb.create_sheet("📅 Monthly Trends")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = "📅  Monthly Attendance Trends"
    c.font      = font(bold=True, color=C_WHITE, size=16)
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="left")
    ws.row_dimensions[1].height = 30

    months = (df.groupby("month_name")["present"]
                .agg(["mean","sum","count"])
                .rename(columns={"mean":"rate","sum":"present","count":"total"}))
    month_order = ["September","October","November","December",
                   "January","February","March","April","May"]
    months = months.reindex([m for m in month_order if m in months.index])
    months["absent"] = months["total"] - months["present"]

    # ── Class-level monthly table ──────────────────────────────────────────────
    section_title(ws, 2, 1, "Class Attendance by Month", merge_to_col=5)
    header_row(ws, 3, range(1,6),
               ["Month","School Days","Present","Absent","Rate"])
    for i, (month, row) in enumerate(months.iterrows()):
        r  = i + 4
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        data_cell(ws, r, 1, month,          bg=bg, h="left", bold=True)
        data_cell(ws, r, 2, int(row.total/len(STUDENTS)), bg=bg)
        data_cell(ws, r, 3, int(row.present), bg=bg, color=C_GREEN)
        data_cell(ws, r, 4, int(row.absent),  bg=bg, color=C_RED)
        data_cell(ws, r, 5, row.rate,        bg=bg, bold=True, number_format="0.0%")

    # Line chart – class rate over months
    chart = LineChart()
    chart.title        = "Class Attendance Rate — Monthly"
    chart.y_axis.title = "Attendance Rate"
    chart.x_axis.title = "Month"
    chart.style        = 10
    chart.height       = 12
    chart.width        = 20

    nrows = len(months)
    data_ref = Reference(ws, min_col=5, max_col=5,
                         min_row=3, max_row=3+nrows)
    cats_ref = Reference(ws, min_col=1, max_col=1,
                         min_row=4, max_row=3+nrows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill    = C_TEAL
    chart.series[0].graphicalProperties.line.width        = 25000
    chart.series[0].marker.symbol                         = "circle"
    ws.add_chart(chart, "G2")

    # ── Per-student monthly heatmap table ─────────────────────────────────────
    heat_row = 3 + nrows + 3
    section_title(ws, heat_row, 1, "Per-Student Monthly Attendance Rate",
                  merge_to_col=len(month_order)+1, bg=C_NAVY)

    header_row(ws, heat_row+1, [1],     ["Student"])
    header_row(ws, heat_row+1, range(2, len(month_order)+2), month_order,
               bg=C_TEAL)

    stu_month = (df.groupby(["student","month_name"])["present"]
                   .mean().unstack(fill_value=0))
    stu_month = stu_month.reindex(
        columns=[m for m in month_order if m in stu_month.columns]
    )

    for i, (stu, row) in enumerate(stu_month.iterrows()):
        r  = heat_row + 2 + i
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        data_cell(ws, r, 1, stu, bg=bg, h="left", bold=True)
        for j, m in enumerate(stu_month.columns):
            v   = row[m]
            mbg = ("2D6A4F" if v >= .95 else
                   "52B788" if v >= .85 else
                   "B7E4C7" if v >= .75 else
                   "FFDD57" if v >= .65 else "E63946")
            mfg = (C_WHITE if v >= .85 or v < .65 else C_DARK)
            data_cell(ws, r, 2+j, v, bg=mbg, bold=False,
                      number_format="0%", color=mfg)

    set_col_widths(ws, {get_column_letter(i): (22 if i==1 else 11)
                        for i in range(1, 14)})


# ── Sheet 4: Day-of-Week Patterns ─────────────────────────────────────────────
def build_dow(wb, df):
    ws = wb.create_sheet("📆 Day-of-Week Patterns")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "📆  Attendance Patterns by Day of Week"
    c.font      = font(bold=True, color=C_WHITE, size=16)
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="left")
    ws.row_dimensions[1].height = 30

    days_order = ["Monday","Tuesday","Wednesday","Thursday","Friday"]
    dow = (df.groupby("weekday")["present"]
             .agg(rate="mean", present="sum", total="count")
             .reindex(days_order))
    dow["absent"] = dow["total"] - dow["present"]

    header_row(ws, 2, range(1,6),
               ["Day","School Days","Present","Absent","Attend. Rate"])
    day_colors = {
        "Monday":    "AED9E0","Tuesday":  "B5EAD7",
        "Wednesday": "FFDAC1","Thursday": "FFB7B2","Friday": "C7CEEA"
    }
    for i, (day, row) in enumerate(dow.iterrows()):
        r = i + 3
        data_cell(ws, r, 1, day,            bg=day_colors[day], h="left", bold=True)
        data_cell(ws, r, 2, int(row.total), bg=C_LIGHT)
        data_cell(ws, r, 3, int(row.present), bg=C_LIGHT, color=C_GREEN)
        data_cell(ws, r, 4, int(row.absent),  bg=C_LIGHT, color=C_RED)
        data_cell(ws, r, 5, row.rate,        bg=C_LIGHT, bold=True, number_format="0.0%")

    # Bar chart
    chart = BarChart()
    chart.title        = "Attendance Rate by Day of Week"
    chart.y_axis.title = "Attendance Rate"
    chart.x_axis.title = "Day"
    chart.style        = 10
    chart.height       = 12
    chart.width        = 18
    chart.grouping     = "clustered"

    data_ref = Reference(ws, min_col=5, max_col=5, min_row=2, max_row=7)
    cats_ref = Reference(ws, min_col=1, max_col=1, min_row=3, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = C_GOLD
    ws.add_chart(chart, "G2")

    # ── Per-student by day ─────────────────────────────────────────────────────
    section_title(ws, 10, 1, "Per-Student Rate by Day of Week",
                  merge_to_col=6, bg=C_NAVY)
    header_row(ws, 11, [1], ["Student"])
    header_row(ws, 11, range(2,7), days_order, bg=C_TEAL)

    stu_dow = (df.groupby(["student","weekday"])["present"]
                 .mean().unstack(fill_value=0))
    stu_dow = stu_dow.reindex(columns=days_order, fill_value=0)

    for i, (stu, row) in enumerate(stu_dow.iterrows()):
        r  = 12 + i
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        data_cell(ws, r, 1, stu, bg=bg, h="left", bold=True)
        for j, day in enumerate(days_order):
            v = row[day]
            dbg = ("D1FAE5" if v >= .90 else
                   "DBEAFE" if v >= .80 else
                   "FEF3C7" if v >= .70 else "FEE2E2")
            data_cell(ws, r, 2+j, v, bg=dbg, number_format="0%")

    set_col_widths(ws, {"A":22,"B":13,"C":12,"D":10,"E":10,"F":13,
                        "G":3,"H":3,"I":3,"J":3,"K":3,"L":20})


# ── Sheet 5: Weekly Tracker ────────────────────────────────────────────────────
def build_weekly(wb, df):
    ws = wb.create_sheet("📈 Weekly Tracker")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = "📈  Weekly Attendance Tracker"
    c.font      = font(bold=True, color=C_WHITE, size=16)
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="left")
    ws.row_dimensions[1].height = 30

    weekly = (df.groupby("week")["present"]
                .agg(rate="mean", total="count")
                .reset_index())
    weekly["absent_rate"] = 1 - weekly["rate"]

    header_row(ws, 2, range(1,4), ["Week #","Attend. Rate","Absence Rate"])
    for i, row in weekly.iterrows():
        r  = i + 3
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        data_cell(ws, r, 1, int(row.week),        bg=bg)
        data_cell(ws, r, 2, row.rate,             bg=bg, number_format="0.0%")
        data_cell(ws, r, 3, row.absent_rate,      bg=bg, number_format="0.0%")

    # Line chart – weekly trend
    chart = LineChart()
    chart.title        = "Weekly Attendance Rate"
    chart.y_axis.title = "Rate"
    chart.x_axis.title = "Week"
    chart.style        = 10
    chart.height       = 14
    chart.width        = 28

    nw = len(weekly)
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=2, max_row=2+nw)
    cats_ref = Reference(ws, min_col=1, max_col=1, min_row=3, max_row=2+nw)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = C_TEAL
    chart.series[0].graphicalProperties.line.width     = 20000
    chart.series[1].graphicalProperties.line.solidFill = C_CORAL
    chart.series[1].graphicalProperties.line.width     = 20000
    ws.add_chart(chart, "E2")

    set_col_widths(ws, {"A":10,"B":14,"C":14})


# ── Sheet 6: At-Risk Students ─────────────────────────────────────────────────
def build_at_risk(wb, df):
    ws = wb.create_sheet("⚠️ At-Risk Students")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "⚠️  At-Risk Student Report  (Attendance < 80%)"
    c.font      = font(bold=True, color=C_WHITE, size=16)
    c.fill      = fill(C_CORAL)
    c.alignment = align(h="left")
    ws.row_dimensions[1].height = 30

    stu = (df.groupby("student")["present"]
             .agg(total="count", present="sum")
             .reset_index())
    stu["rate"]   = stu["present"] / stu["total"]
    stu["absent"] = stu["total"] - stu["present"]
    at_risk       = stu[stu["rate"] < 0.80].sort_values("rate")

    header_row(ws, 2, range(1,8),
               ["Student","Days Present","Days Absent","Total Days",
                "Attend. Rate","Days to 80%","Risk Level"], bg=C_CORAL)

    for i, row in at_risk.iterrows():
        r  = at_risk.index.get_loc(i) + 3
        bg = C_LIGHT if r % 2 == 0 else C_WHITE

        days_needed = max(0, int(0.80 * row.total - row.present))
        risk = ("🔴 Critical" if row.rate < 0.60 else
                "🟠 High"    if row.rate < 0.70 else "🟡 Moderate")
        rbg  = ("FEE2E2" if row.rate < 0.60 else
                "FEF3C7" if row.rate < 0.70 else "FFFBEB")

        data_cell(ws, r, 1, row.student,          bg=bg, h="left", bold=True)
        data_cell(ws, r, 2, int(row.present),     bg=bg, color=C_GREEN)
        data_cell(ws, r, 3, int(row.absent),      bg=bg, color=C_RED)
        data_cell(ws, r, 4, int(row.total),       bg=bg)
        data_cell(ws, r, 5, row.rate,             bg=rbg, bold=True, number_format="0.0%", color=C_RED)
        data_cell(ws, r, 6, days_needed,          bg=bg, color=C_CORAL)
        data_cell(ws, r, 7, risk,                 bg=rbg, bold=True)

    # Summary box
    sr = len(at_risk) + 4
    ws.merge_cells(f"A{sr}:G{sr}")
    c = ws[f"A{sr}"]
    c.value     = (f"📌  {len(at_risk)} student(s) are below the 80% attendance threshold. "
                   "Immediate outreach recommended.")
    c.font      = font(bold=True, color=C_WHITE, size=12)
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="left")

    set_col_widths(ws, {"A":22,"B":14,"C":13,"D":12,"E":13,"F":13,"G":13})


# ── Sheet 7: Dashboard ────────────────────────────────────────────────────────
def build_dashboard(wb, df):
    ws = wb.create_sheet("🏠 Dashboard")
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:L2")
    c = ws["A1"]
    c.value     = "🏫  STUDENT ATTENDANCE ANALYTICS DASHBOARD  —  AY 2024–25"
    c.font      = font(bold=True, color=C_WHITE, size=20)
    c.fill      = fill(C_NAVY)
    c.alignment = align()
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    overall_rate  = df["present"].mean()
    total_stu     = df["student"].nunique()
    total_days    = df["date"].nunique()
    at_risk_count = (df.groupby("student")["present"].mean() < 0.80).sum()

    kpis = [
        ("Overall Rate",      f"{overall_rate:.1%}", C_TEAL),
        ("Total Students",    str(total_stu),          C_NAVY),
        ("School Days",       str(total_days),          C_GREEN),
        ("At-Risk Students",  str(at_risk_count),       C_CORAL),
    ]

    for k, (label, value, color) in enumerate(kpis):
        col = k * 3 + 1
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+2)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+2)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+2)
        c4 = ws.cell(row=4, column=col, value=label)
        c4.font = font(bold=True, color=C_WHITE, size=11)
        c4.fill = fill(color); c4.alignment = align()
        c5 = ws.cell(row=5, column=col, value=value)
        c5.font = font(bold=True, color=color, size=26)
        c5.fill = fill(C_LIGHT); c5.alignment = align()
        c6 = ws.cell(row=6, column=col, value=" ")
        c6.fill = fill(color)

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 40
    ws.row_dimensions[6].height = 6

    # Tier breakdown table
    section_title(ws, 8, 1, "Attendance by Tier", merge_to_col=4, bg=C_TEAL)
    header_row(ws, 9, range(1,5), ["Tier","Students","Avg Rate","Risk"], bg=C_TEAL)
    tier_data = (df.groupby("tier")
                   .agg(students=("student","nunique"), rate=("present","mean"))
                   .reset_index())
    tier_colors2 = {"excellent":"065F46","good":"1E40AF","average":"92400E","poor":"991B1B"}
    tier_bgs    = {"excellent":"D1FAE5","good":"DBEAFE","average":"FEF3C7","poor":"FEE2E2"}
    tier_order  = ["excellent","good","average","poor"]
    tier_data["order"] = tier_data["tier"].map({t:i for i,t in enumerate(tier_order)})
    tier_data = tier_data.sort_values("order")
    for i, row in tier_data.iterrows():
        r  = 10 + list(tier_data.index).index(i)
        bg = tier_bgs.get(row.tier, C_WHITE)
        fg = tier_colors2.get(row.tier, C_DARK)
        data_cell(ws, r, 1, row.tier.capitalize(), bg=bg, bold=True, color=fg)
        data_cell(ws, r, 2, int(row.students),     bg=C_LIGHT)
        data_cell(ws, r, 3, row.rate,              bg=bg, bold=True, number_format="0.0%", color=fg)
        risk = ("Low" if row.rate >= .90 else "Medium" if row.rate >= .75 else "High")
        data_cell(ws, r, 4, risk, bg=bg, color=fg, bold=True)

    set_col_widths(ws, {get_column_letter(i): 14 for i in range(1, 13)})
    ws.column_dimensions["A"].width = 22

    # Note
    note_row = 16
    ws.merge_cells(f"A{note_row}:L{note_row}")
    c = ws[f"A{note_row}"]
    c.value     = ("💡  Navigate using the sheet tabs below  |  "
                   "📋 Attendance Log  |  📊 Student Summary  |  "
                   "📅 Monthly Trends  |  📆 Day-of-Week  |  📈 Weekly Tracker  |  ⚠️ At-Risk")
    c.font      = font(italic=True, color=C_GRAY, size=10)
    c.fill      = fill(C_LIGHT)
    c.alignment = align(h="left")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print("⏳  Generating attendance data …")
    days = school_days(SCHOOL_START, SCHOOL_END)
    df   = generate_attendance(STUDENTS, days)
    print(f"    {len(df):,} records  |  {len(days)} school days  |  {len(STUDENTS)} students")

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    print("📊  Building workbook sheets …")
    build_dashboard(wb, df)
    build_summary(wb, df)
    build_monthly(wb, df)
    build_dow(wb, df)
    build_weekly(wb, df)
    build_raw_log(wb, df)
    build_at_risk(wb, df)

    out = "Student_Attendance_Analysis.xlsx"
    wb.save(out)
    print(f"✅  Saved → {out}")


if __name__ == "__main__":
    main()
